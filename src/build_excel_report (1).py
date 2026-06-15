"""
Ridgeline Foods — Freight Accrual Excel Report Generator
Produces a multi-tab Excel workbook:
  1. Accrual Summary (journal entry ready)
  2. Shipment Detail (one row per shipment with full charge breakdown)
  3. Data Quality Flags
  4. Variance Analysis vs Denise's trailing average
  5. Assumptions & Controls
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from accrual_engine import run_accrual_engine

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from datetime import date

# ── Colors (standard finance model palette) ──────────────────────────────────
C_HEADER_BG  = "1F3864"   # dark navy
C_HEADER_FG  = "FFFFFF"   # white
C_SUBHDR_BG  = "D6E4F0"   # light blue
C_PEAK       = "E8F5E9"   # light green
C_HEARTLAND  = "FFF3E0"   # light orange
C_COASTAL    = "E3F2FD"   # light blue
C_TOTAL_BG   = "FFF9C4"   # light yellow
C_FLAG_BG    = "FFEBEE"   # light red
C_INPUT_FONT = "0000FF"   # blue for hardcoded inputs
C_FORMULA    = "000000"   # black for formulas
C_WARN       = "FF6600"   # orange for warnings
C_GOOD       = "006400"   # dark green for ok

CARRIER_COLORS = {
    "Peak Logistics": C_PEAK,
    "Heartland Freight": C_HEARTLAND,
    "Coastal Express": C_COASTAL,
}

def make_border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr(ws, row, col, value, bg=C_HEADER_BG, fg=C_HEADER_FG, bold=True, center=True):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, color=fg, size=10)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center" if center else "left", vertical="center", wrap_text=True)
    cell.border = make_border()
    return cell

def data_cell(ws, row, col, value, fmt=None, bold=False, bg=None, font_color=C_FORMULA, left=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(name="Arial", bold=bold, size=10, color=font_color)
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    if fmt:
        cell.number_format = fmt
    cell.alignment = Alignment(horizontal="left" if left else "general", vertical="center")
    cell.border = make_border()
    return cell

def currency(ws, row, col, value, bold=False, bg=None, font_color=C_FORMULA):
    c = data_cell(ws, row, col, value, fmt='$#,##0.00;($#,##0.00);"-"', bold=bold, bg=bg, font_color=font_color)
    c.alignment = Alignment(horizontal="right", vertical="center")
    return c

def pct(ws, row, col, value, bold=False, bg=None):
    c = data_cell(ws, row, col, value, fmt='0.0%;(0.0%);"-"', bold=bold, bg=bg)
    c.alignment = Alignment(horizontal="right", vertical="center")
    return c

# ─────────────────────────────────────────────────────────────────────────────
# TAB 1: ACCRUAL SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def build_summary_tab(ws, result):
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 20

    # Title block
    ws.merge_cells("A1:G1")
    t = ws["A1"]
    t.value = "RIDGELINE FOODS — APRIL 2026 FREIGHT ACCRUAL"
    t.font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:G2")
    s = ws["A2"]
    s.value = f"Prepared: {date.today().strftime('%B %d, %Y')}   |   Service Month: April 2026   |   Estimates based on 3PL shipment log + contracted rate cards"
    s.font = Font(name="Arial", size=9, italic=True)
    s.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # Column headers row 4
    r = 4
    hdr(ws, r, 1, "Carrier")
    hdr(ws, r, 2, "Shipments")
    hdr(ws, r, 3, "Base Charges")
    hdr(ws, r, 4, "Fuel Surcharge")
    hdr(ws, r, 5, "Accessorials")
    hdr(ws, r, 6, "Net Accrual $")
    hdr(ws, r, 7, "Notes")
    ws.row_dimensions[r].height = 24

    carriers = ["Peak Logistics", "Heartland Freight", "Coastal Express"]
    carrier_rows = {}
    for i, carrier in enumerate(carriers):
        s = result["summary"][carrier]
        row = 5 + i
        carrier_rows[carrier] = row
        bg = CARRIER_COLORS[carrier]
        data_cell(ws, row, 1, carrier, bold=True, bg=bg, left=True)
        data_cell(ws, row, 2, s["shipment_count"], bg=bg)
        currency(ws, row, 3, s["base_total"], bg=bg)
        currency(ws, row, 4, s["fuel_total"], bg=bg)
        currency(ws, row, 5, s["accessorial_total"], bg=bg)
        currency(ws, row, 6, s["net_accrual"], bold=True, bg=bg)
        notes = {
            "Peak Logistics": "Per-mile rates + 14% FSC. Mileage from rate card & invoice back-calc.",
            "Heartland Freight": "Flat zone rates. FSC included. Apr = Q2 wk 1 → Tier 1 (0% disc).",
            "Coastal Express": "Per-lb rates by region (ZIP). 9.5% FSC fixed thru Q2 2026.",
        }
        data_cell(ws, row, 7, notes[carrier], left=True, bg=bg)
        ws.row_dimensions[row].height = 20

    # Total row
    total_row = 8
    data_cell(ws, total_row, 1, "TOTAL", bold=True, bg=C_TOTAL_BG)
    data_cell(ws, total_row, 2, f'=SUM(B5:B7)', bg=C_TOTAL_BG)
    for col in range(3, 7):
        cl = get_column_letter(col)
        currency(ws, total_row, col, f'=SUM({cl}5:{cl}7)', bold=True, bg=C_TOTAL_BG)
    data_cell(ws, total_row, 7, "← Book to Freight Expense / Accrued Liabilities", bold=True, bg=C_TOTAL_BG, left=True)
    ws.row_dimensions[total_row].height = 22

    # Adjustment buffer note
    ws["A10"] = "Adjustment buffer:"
    ws["A10"].font = Font(name="Arial", size=9, italic=True)
    ws["B10"] = f"{result['adjustment_rate']*100:.2f}% of gross"
    ws["B10"].font = Font(name="Arial", size=9, italic=True)
    ws["C10"] = "Historical credit/dispute adjustments avg. Applied as conservative haircut on gross estimate."
    ws.merge_cells("C10:G10")
    ws["C10"].font = Font(name="Arial", size=9, italic=True)

    # Journal Entry block
    r = 12
    ws.merge_cells(f"A{r}:G{r}")
    je = ws[f"A{r}"]
    je.value = "SUGGESTED JOURNAL ENTRY"
    je.font = Font(name="Arial", bold=True, size=11)
    je.fill = PatternFill("solid", fgColor=C_SUBHDR_BG)
    je.alignment = Alignment(horizontal="left")

    r += 1
    for col, label in enumerate(["Account", "Description", "Debit", "Credit", "Department", "Memo"], 1):
        hdr(ws, r, col, label, bg="4472C4")
    r += 1
    # Debit freight expense
    data_cell(ws, r, 1, "6300 - Freight Expense", left=True)
    data_cell(ws, r, 2, "April 2026 outbound freight accrual", left=True)
    currency(ws, r, 3, f"=F8")   # total net accrual
    currency(ws, r, 4, None)
    data_cell(ws, r, 5, "Operations", left=True)
    data_cell(ws, r, 6, "Accrual — invoices expected May 8–12", left=True)
    r += 1
    # Credit accrued liabilities
    data_cell(ws, r, 1, "2100 - Accrued Liabilities", left=True)
    data_cell(ws, r, 2, "Freight carriers — April 2026", left=True)
    currency(ws, r, 3, None)
    currency(ws, r, 4, f"=F8")
    data_cell(ws, r, 5, "", left=True)
    data_cell(ws, r, 6, "To reverse upon receipt of invoices", left=True)

    # Denise comparison block
    r += 2
    ws.merge_cells(f"A{r}:G{r}")
    dc = ws[f"A{r}"]
    dc.value = "COMPARISON: THIS METHOD vs DENISE'S TRAILING 3-MONTH AVERAGE"
    dc.font = Font(name="Arial", bold=True, size=11)
    dc.fill = PatternFill("solid", fgColor=C_SUBHDR_BG)

    r += 1
    for col, label in enumerate(["Carrier", "Denise 3-Mo Avg", "This Estimate", "$ Difference", "% Difference", "Method"], 1):
        hdr(ws, r, col, label, bg="4472C4")

    denise_avgs = {
        "Peak Logistics": 34406,
        "Heartland Freight": 35333,
        "Coastal Express": 25533,
    }
    start_r = r + 1
    for i, carrier in enumerate(carriers):
        row = start_r + i
        bg = CARRIER_COLORS[carrier]
        data_cell(ws, row, 1, carrier, bold=True, bg=bg, left=True)
        currency(ws, row, 2, denise_avgs[carrier], bg=bg, font_color=C_INPUT_FONT)
        currency(ws, row, 3, result["summary"][carrier]["net_accrual"], bg=bg)
        currency(ws, row, 4, f"=C{row}-B{row}", bg=bg)
        pct(ws, row, 5, f"=(C{row}-B{row})/B{row}", bg=bg)
        methods = {
            "Peak Logistics": "Rate card: per-mile × weight tier × miles + 14% FSC + accessorials",
            "Heartland Freight": "Rate card: zone flat rate + Q2 Tier 1 (0% disc) + accessorials",
            "Coastal Express": "Rate card: per-lb × region + 9.5% FSC + residential surcharges",
        }
        data_cell(ws, row, 6, methods[carrier], left=True, bg=bg)

    total_denise = sum(denise_avgs.values())
    total_row2 = start_r + 3
    data_cell(ws, total_row2, 1, "TOTAL", bold=True, bg=C_TOTAL_BG)
    currency(ws, total_row2, 2, total_denise, bold=True, bg=C_TOTAL_BG, font_color=C_INPUT_FONT)
    currency(ws, total_row2, 3, f"=SUM(C{start_r}:C{start_r+2})", bold=True, bg=C_TOTAL_BG)
    currency(ws, total_row2, 4, f"=C{total_row2}-B{total_row2}", bold=True, bg=C_TOTAL_BG)
    pct(ws, total_row2, 5, f"=(C{total_row2}-B{total_row2})/B{total_row2}", bold=True, bg=C_TOTAL_BG)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2: SHIPMENT DETAIL
# ─────────────────────────────────────────────────────────────────────────────

def build_detail_tab(ws, result):
    cols = [
        ("Shipment ID", 14), ("Carrier", 18), ("Destination City", 16),
        ("ST", 5), ("ZIP", 8), ("Weight (lbs)", 12), ("Zone / Region", 22),
        ("Base Charge", 12), ("Fuel Surcharge", 14), ("Accessorials", 12),
        ("Accessorial Detail", 28), ("Total Estimate", 14), ("Notes / Flags", 35),
    ]
    for i, (label, width) in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
        hdr(ws, 1, i, label)

    ws.row_dimensions[1].height = 22

    for row_idx, est in enumerate(result["estimates"], 2):
        bg = CARRIER_COLORS.get(est.carrier, "FFFFFF")
        flag_str = "; ".join(est.flags) if est.flags else ""
        note_bg = "FFCCCC" if est.flags else bg  # red tint if flagged

        def dc(col, val, fmt=None, left=False, bold=False):
            c = ws.cell(row=row_idx, column=col, value=val)
            c.font = Font(name="Arial", size=9, bold=bold)
            c.fill = PatternFill("solid", fgColor=note_bg if col == 13 else bg)
            c.border = make_border()
            c.alignment = Alignment(horizontal="left" if left else "right" if fmt else "center", vertical="center")
            if fmt:
                c.number_format = fmt
            return c

        dc(1, est.shipment_id, left=True)
        dc(2, est.carrier, left=True)
        dc(3, est.dest_city, left=True)
        dc(4, est.dest_state)
        dc(5, est.dest_zip, left=True)
        dc(6, round(est.weight, 1), fmt='#,##0.0')
        dc(7, est.zone_or_region, left=True)
        dc(8, est.base_charge, fmt='$#,##0.00')
        dc(9, est.fuel_surcharge, fmt='$#,##0.00')
        dc(10, est.accessorial_total, fmt='$#,##0.00')
        dc(11, est.accessorial_detail, left=True)
        dc(12, est.total_charge, fmt='$#,##0.00', bold=True)
        dc(13, flag_str if flag_str else est.notes, left=True)

    # Totals row
    total_row = len(result["estimates"]) + 2
    ws.cell(row=total_row, column=1, value="TOTALS").font = Font(name="Arial", bold=True)
    ws.cell(row=total_row, column=1).fill = PatternFill("solid", fgColor=C_TOTAL_BG)
    for col in [6, 8, 9, 10, 12]:
        cl = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({cl}2:{cl}{total_row-1})")
        c.font = Font(name="Arial", bold=True)
        c.fill = PatternFill("solid", fgColor=C_TOTAL_BG)
        c.number_format = '$#,##0.00'
        c.border = make_border()

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(13)}1"


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3: DATA QUALITY FLAGS
# ─────────────────────────────────────────────────────────────────────────────

def build_flags_tab(ws, result):
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 65

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "DATA QUALITY FLAGS — Review before posting accrual"
    t.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor="C62828")
    t.alignment = Alignment(horizontal="center")

    for col, label in enumerate(["Shipment ID", "Carrier", "Flag / Issue"], 1):
        hdr(ws, 2, col, label, bg="B71C1C")

    flags = result["flags"]
    if not flags:
        ws.cell(row=3, column=1, value="✓ No data quality issues found").font = Font(name="Arial", color=C_GOOD)
    else:
        for i, flag in enumerate(flags, 3):
            ws.cell(row=i, column=1, value=flag.get("shipment", "")).font = Font(name="Arial", size=9)
            ws.cell(row=i, column=2, value=flag.get("carrier", "")).font = Font(name="Arial", size=9)
            ws.cell(row=i, column=3, value=flag.get("flag", "")).font = Font(name="Arial", size=9)
            for col in range(1, 4):
                ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=C_FLAG_BG)
                ws.cell(row=i, column=col).border = make_border()

    ws.row_dimensions[1].height = 24


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4: ASSUMPTIONS & CONTROLS
# ─────────────────────────────────────────────────────────────────────────────

def build_assumptions_tab(ws, adjustment_rate):
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 40

    ws.merge_cells("A1:C1")
    t = ws["A1"]
    t.value = "ASSUMPTIONS, CONTROLS & METHODOLOGY"
    t.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    t.fill = PatternFill("solid", fgColor=C_HEADER_BG)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    sections = [
        ("PEAK LOGISTICS", [
            ("Rate basis", "Per-mile × weight-tier rate. Weight tiers: <500 lbs=$3.25/mi, 500–2000=$4.80/mi, >2000=$6.40/mi"),
            ("Fuel surcharge", "14% of base freight charge. Contractually fixed through Q2 2026."),
            ("Minimum charge", "$185.00 per shipment regardless of distance or weight."),
            ("Mileage source", "Rate card verified for 11 cities. Remaining cities derived from historical invoice back-calculation (base÷rate=implied miles)."),
            ("SLC-origin shipments", "6 April Peak Logistics shipments originate from Salt Lake City (one is the duplicate SHP-10033, excluded — 5 remain in this accrual). Mileages estimated using Denver-routing as a proxy. These are unusual backhaul routings — flagged for CFO review."),
            ("Out-of-territory", "SHP-10006 (Reno, NV) is outside Peak's Mountain West territory. Estimated at 900-mile proxy; contact Peak for actual quote."),
        ]),
        ("HEARTLAND FREIGHT", [
            ("Rate basis", "Flat zone rate per shipment (Zone 1–4). Fuel surcharge INCLUDED in flat rate — do not add separate FSC."),
            ("Zone assignment", "Determined by first 3 digits of destination ZIP code. ZIP prefix table from rate card applied."),
            ("Volume discount tier", "April = Q2 Week 1. Quarter resets April 1. All April shipments fall in Tier 1 (0% discount). NOTE: If Heartland confirms Ridgeline carried forward a prior tier from a negotiated rollover, revise discount assumption."),
            ("Quarter reset impact", "January 2026 actual was $38,200 vs Denise's $31,067 estimate — a $7,133 miss driven by Q1 reset. Same risk applies to April (Q2 reset). This model explicitly accounts for Tier 1."),
        ]),
        ("COASTAL EXPRESS", [
            ("Rate basis", "Per-lb rates by region: SoCal (ZIP 90000–92899) $0.48/lb, NorCal (93000–96199) $0.55/lb, PNW (97000–99499) $0.72/lb."),
            ("Minimum charge", "$28.00 per shipment base. Fuel surcharge and accessorials add on top."),
            ("Fuel surcharge", "9.5% of base charge. Contractually fixed through Q2 2026. NOT applied to accessorial fees."),
            ("Residential surcharge", "Tiered: <50 lbs=$12.50, 50–500 lbs=$35.00, >500 lbs=$65.00. Applied to all shipments with residential=TRUE in 3PL log."),
        ]),
        ("GENERAL CONTROLS", [
            ("Carrier name normalization", "3PL log contains 9 carrier name variants across 3 carriers. All mapped to canonical names before rate application."),
            ("Duplicate detection", "SHP-10033 appeared twice in April log (identical record). Second occurrence excluded."),
            ("Missing weight imputation", "5 shipments had blank weight_lbs. Imputed using carrier-specific avg lbs/unit from remaining April records."),
            ("Adjustment buffer", f"{adjustment_rate*100:.2f}% applied to gross estimate per carrier. Derived from 6 months of actual invoice adjustments (22 adjustments, all negative) as a % of pre-adjustment gross — reflects historical pattern of carrier credit adjustments."),
            ("Accrual reversal", "Book as debit Freight Expense / credit Accrued Liabilities. Reverse upon invoice receipt (expected May 8–12). If invoice differs >5% from accrual, investigate before reversing."),
            ("Confidence range", "Estimates based on April 3PL data + contracted rates. Actual invoices may vary due to: carrier weight reclassification, additional accessorials not in 3PL log, or billing disputes."),
        ]),
    ]

    row = 3
    for section_title, items in sections:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws[f"A{row}"]
        c.value = section_title
        c.font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="4472C4")
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 18
        row += 1

        # Column header
        for col, label in enumerate(["Assumption / Control", "Detail", "Source / Rationale"], 1):
            hdr(ws, row, col, label, bg=C_SUBHDR_BG, fg="000000", bold=True)
        ws.row_dimensions[row].height = 16
        row += 1

        for key, val in items:
            ws.cell(row=row, column=1, value=key).font = Font(name="Arial", bold=True, size=9)
            c = ws.cell(row=row, column=2, value=val)
            c.font = Font(name="Arial", size=9)
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=row, column=3, value="Rate card / invoice analysis").font = Font(name="Arial", size=9, italic=True)
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = make_border()
            ws.row_dimensions[row].height = 30
            row += 1
        row += 1


# ─────────────────────────────────────────────────────────────────────────────
# MAIN: BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────

def build_workbook(shipments_file: str, output_path: str, invoices_file: str = None):
    print(f"Running accrual engine on {shipments_file}...")
    result = run_accrual_engine(shipments_file, invoices_file)

    wb = Workbook()

    # Tab 1: Summary
    ws_sum = wb.active
    ws_sum.title = "Accrual Summary"
    build_summary_tab(ws_sum, result)
    print(f"  ✓ Summary tab built")

    # Tab 2: Detail
    ws_det = wb.create_sheet("Shipment Detail")
    build_detail_tab(ws_det, result)
    print(f"  ✓ Detail tab built ({len(result['estimates'])} rows)")

    # Tab 3: Flags
    ws_flags = wb.create_sheet("Data Quality Flags")
    build_flags_tab(ws_flags, result)
    print(f"  ✓ Flags tab built ({len(result['flags'])} flags)")

    # Tab 4: Assumptions
    ws_assum = wb.create_sheet("Assumptions & Controls")
    build_assumptions_tab(ws_assum, result["adjustment_rate"])
    print(f"  ✓ Assumptions tab built")

    wb.save(output_path)
    print(f"\n✅ Workbook saved: {output_path}")

    # Print key numbers
    print(f"\n{'='*50}")
    print(f"April 2026 Freight Accrual")
    for carrier, s in result["summary"].items():
        print(f"  {carrier}: ${s['net_accrual']:,.2f}")
    print(f"  TOTAL: ${result['total_accrual']:,.2f}")
    print(f"{'='*50}")

    return result


if __name__ == "__main__":
    HERE = os.path.dirname(os.path.abspath(__file__))
    DATA_DIR = os.path.join(os.path.dirname(HERE), "data")
    OUT_DIR = os.path.join(os.path.dirname(HERE), "output")
    os.makedirs(OUT_DIR, exist_ok=True)
    build_workbook(
        os.path.join(DATA_DIR, "shipments_apr2026.csv"),
        os.path.join(OUT_DIR, "ridgeline_freight_accrual_apr2026.xlsx"),
        os.path.join(DATA_DIR, "freight_invoices_oct2025_mar2026_v2.csv"),
    )
